from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Any

def generate_excel_report(title: str, headers: List[str], data: List[Dict[str, Any]], keys: List[str]):
    """
    Genera un archivo Excel dinámico a partir de una lista de diccionarios.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Título en la primera fila
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title.upper())
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = center_alignment

    # Cabeceras
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Datos
    for i, row_data in enumerate(data, 4):
        for j, key in enumerate(keys, 1):
            val = row_data.get(key, "")
            # Formatear números
            if isinstance(val, (int, float)):
                cell = ws.cell(row=i, column=j, value=float(val))
                if "total" in key or "monto" in key or "subtotal" in key or "iva" in key:
                    cell.number_format = '"$"#,##0.00'
            else:
                cell = ws.cell(row=i, column=j, value=str(val))
            
            cell.border = border

    # Auto-ajuste de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
